"""Import immutable pre-Ledger final reports as frozen current-model snapshots.

The source workbooks are business evidence, not new calculation inputs.  They are therefore
read-only, content-addressed, and never inserted into an active workspace slot.  Each row is
mapped deterministically to the current statement node ids and checked back to the workbook's
reported profit before it can be recorded or closed.
"""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import sqlite3
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .model.schema import Model, StatementNode
from .version import engine_version
from .workspace import CLOSED, PeriodState, Workspace, WorkspaceError


CUTOFF = "2026-05"
OPERATOR = "历史账迁移"

# Exact business identities.  Longest-substring filename matching must not be reused here:
# ``皇莉诗旗舰店`` is the Taobao store in these reports while the same phrase is also a JD
# source-file alias.  Historical result ownership is a separate, explicit contract.
LEGACY_STORE_IDS = {
    "PDD乐趣": "pdd_mt9so37w",
    "PDD京禧": "pdd_jingxi",
    "PDD厂家": "legacy_pdd_changjia",
    "PDD喜必顺": "pdd_xibishun",
    "PDD喜舒曼": "pdd_xishuman",
    "PDD嘻嘻": "pdd_xixihaha",
    "PDD国风": "pdd_mt9sojk5",
    "PDD好日子": "pdd_msy75gx9",
    "PDD婚庆": "pdd_hunqing",
    "PDD家居": "legacy_pdd_jiaju",
    "PDD小红薯": "pdd_xiaohongshu_qiqiu",
    "PDD快乐": "pdd_kuailejieqing",
    "PDD意大利": "pdd_yidali",
    "PDD棱智": "legacy_pdd_lingzhi",
    "PDD气球": "pdd_qiqiugongfang",
    "PDD永箔": "legacy_pdd_yongbo",
    "PDD泡泡堂": "pdd_mt9sliee",
    "PDD派对": "legacy_pdd_paidui",
    "PDD浅花涧": "pdd_qianhuajian",
    "PDD皇莉诗": "pdd_huanglishi",
    "PDDzlvoey": "pdd_zlvoey",
    "zlvoey旗舰店": "taobao_zlvoey",
    "京东皇莉诗": "jd_huanglishi",
    "喜必顺旗舰店": "taobao_xibishun",
    "喜舒曼旗舰店": "taobao_mt9s7n9x",
    "喜舒曼礼品旗舰店": "taobao_mt9s7n9x",
    "喜菲喵旗舰店": "taobao_xifeimiao",
    "天楼阁宇": "taobao_tianlougeyu",
    "淘宝天楼阁宇": "taobao_tianlougeyu",
    "家居旗舰店": "taobao_xingyunyuanwang",
    "思福喜旗舰店": "taobao_mt9sluo7",
    "抖店喜品": "douyin_mt9sbkne",
    "抖店喜必顺": "douyin_xibishun",
    "抖店浅花涧": "douyin_qianhuajian",
    "抖店烘焙": "douyin_mszr2dhn",
    "拾梦小屋": "taobao_mt9sc10p",
    "星泽1688": "alibaba1688_xingze",
    "朗歆1688": "alibaba1688_mt2r23jf",
    "浅花涧旗舰店": "taobao_mt9scjag",
    "淘宝喜气洋洋": "taobao_mt9sjzvc",
    "淘宝嘻嘻哈哈": "taobao_mt9scbh6",
    "甜嘴熊旗舰店": "taobao_mt9snd3i",
    "皇莉诗旗舰店": "taobao_msy387nx",
    "美食专家": "taobao_mt8egr48",
    "藏依格": "taobao_mt9smlcz",
    "阳光朵朵旗舰店": "taobao_mt9smaj1",
}

COMPONENT_TO_NODE = {
    "交易收款": "n_receipt",
    "交易退款": "n_refund",
    "软件服务费": "n_software",
    "交易赔付": "n_compensation",
    "营销费用": "n_marketing",
    "发货运费": "n_freight",
    "订单成本": "n_goods",
    "补发成本": "n_reshipment",
    "本金佣金": "n_brushing",
    "代购代发": "n_dropship",
    "广告费": "n_ad",
}

ZERO_NODES = {
    "n_expense", "n_logistics", "n_cross_border", "n_insurance", "n_small_payment",
}


def _now() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def _money(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"金额看不懂：{value!r}") from exc


def _hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _period(value: Any) -> str:
    text = str(value or "").strip()
    match = re.fullmatch(r"(20\d{2})[-/]?(\d{2})", text)
    if not match:
        raise ValueError(f"年月看不懂：{value!r}")
    period = f"{match.group(1)}-{match.group(2)}"
    if not 1 <= int(match.group(2)) <= 12:
        raise ValueError(f"月份超出范围：{value!r}")
    return period


_DATED_SUFFIX = re.compile(
    r"(?:20)?\d{1,2}[.\-/]?\d{1,2}\s*[-至]\s*(?:20)?\d{0,2}[.\-/]?\d{1,2}.*$"
)


def canonical_legacy_name(value: Any) -> str:
    name = str(value or "").strip()
    if not name or name.startswith("#"):
        return ""
    name = name.replace("（", "(").replace("）", ")")
    # April 2026 contains split-period rows such as 浅花涧旗舰店4.1-4.14.  They are
    # pieces of the same store and must be aggregated, not invented as extra stores.
    name = _DATED_SUFFIX.sub("", name).strip()
    return name


@dataclass(slots=True)
class LegacyCell:
    source: Path
    source_sha256: str
    sheet: str
    row_numbers: list[int]
    legacy_names: set[str]
    store_id: str
    period: str
    values: dict[str, Decimal] = field(default_factory=dict)

    def add(self, row_number: int, legacy_name: str, values: dict[str, Decimal]) -> None:
        self.row_numbers.append(row_number)
        self.legacy_names.add(legacy_name)
        for key, value in values.items():
            self.values[key] = self.values.get(key, Decimal("0")) + value


@dataclass(slots=True)
class ImportPlan:
    cells: list[LegacyCell]
    sources: dict[Path, str]
    errors: list[str]
    warnings: list[str]

    @property
    def periods(self) -> list[str]:
        return sorted({cell.period for cell in self.cells})


def _row_values(headers: dict[str, int], row: tuple[Any, ...]) -> dict[str, Decimal]:
    values: dict[str, Decimal] = {}
    for column in [*COMPONENT_TO_NODE, "店铺毛利", "店铺利润"]:
        index = headers.get(column)
        values[column] = _money(row[index] if index is not None and index < len(row) else None)
    return values


def _resolve_name(primary: Any, fallback: Any) -> tuple[str, str]:
    raw = canonical_legacy_name(primary) or canonical_legacy_name(fallback)
    store_id = LEGACY_STORE_IDS.get(raw, "")
    return raw, store_id


def _append(
    grouped: dict[tuple[str, str], LegacyCell],
    *,
    source: Path,
    sha: str,
    sheet: str,
    row_number: int,
    period: str,
    primary_name: Any,
    fallback_name: Any,
    values: dict[str, Decimal],
    errors: list[str],
) -> None:
    legacy_name, store_id = _resolve_name(primary_name, fallback_name)
    if not store_id:
        errors.append(f"{source.name}/{sheet} 第{row_number}行：店铺无法对应：{legacy_name or primary_name!r}")
        return
    key = (store_id, period)
    cell = grouped.get(key)
    if cell is None:
        cell = LegacyCell(source, sha, sheet, [], set(), store_id, period)
        grouped[key] = cell
    elif cell.source != source:
        errors.append(f"{store_id} {period} 同时来自两份终态文件")
        return
    cell.add(row_number, legacy_name, values)


def read_2025_annual(path: Path, grouped: dict, errors: list[str]) -> str:
    sha = _hash(path)
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        sheet = workbook["店铺明细"]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)
        headers = {str(value): index for index, value in enumerate(header) if value is not None}
        required = {"Name", "年月", "交易收款", "店铺利润"}
        if missing := sorted(required - headers.keys()):
            raise ValueError(f"{path.name} 缺列：{'、'.join(missing)}")
        for row_number, row in enumerate(rows, 2):
            if not row[headers["Name"]]:
                continue
            period = _period(row[headers["年月"]])
            if period > CUTOFF:
                errors.append(f"{path.name}/{sheet.title} 第{row_number}行越过截止期：{period}")
                continue
            _append(
                grouped, source=path, sha=sha, sheet=sheet.title, row_number=row_number,
                period=period, primary_name=row[headers["Name"]], fallback_name="",
                values=_row_values(headers, row), errors=errors,
            )
    finally:
        workbook.close()
    return sha


def read_2026_months(path: Path, grouped: dict, errors: list[str]) -> str:
    sha = _hash(path)
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        for month in range(1, 6):
            title = f"{month}月"
            if title not in workbook.sheetnames:
                errors.append(f"{path.name} 缺工作表：{title}")
                continue
            sheet = workbook[title]
            rows = sheet.iter_rows(values_only=True)
            header = next(rows)
            headers = {str(value): index for index, value in enumerate(header) if value is not None}
            required = {"Name", "交易收款", "店铺利润"}
            if missing := sorted(required - headers.keys()):
                errors.append(f"{path.name}/{title} 缺列：{'、'.join(missing)}")
                continue
            period = f"2026-{month:02d}"
            for row_number, row in enumerate(rows, 2):
                fallback = row[headers["Name"]] if headers["Name"] < len(row) else None
                if not fallback:
                    continue
                primary_index = headers.get("店铺名称")
                primary = row[primary_index] if primary_index is not None and primary_index < len(row) else None
                _append(
                    grouped, source=path, sha=sha, sheet=title, row_number=row_number,
                    period=period, primary_name=primary, fallback_name=fallback,
                    values=_row_values(headers, row), errors=errors,
                )
    finally:
        workbook.close()
    return sha


def build_plan(source_2025: Path, source_2026: Path, model: Model) -> ImportPlan:
    grouped: dict[tuple[str, str], LegacyCell] = {}
    errors: list[str] = []
    sources: dict[Path, str] = {}
    for path in (source_2025, source_2026):
        if not path.is_file():
            errors.append(f"历史终态文件不存在：{path}")
    if errors:
        return ImportPlan([], sources, errors, [])
    sources[source_2025] = read_2025_annual(source_2025, grouped, errors)
    sources[source_2026] = read_2026_months(source_2026, grouped, errors)
    known = {store.id for store in model.stores}
    warnings: list[str] = []
    for cell in grouped.values():
        if cell.store_id not in known:
            errors.append(f"当前模型没有历史店铺 {cell.store_id}（{sorted(cell.legacy_names)}）")
        expected = sum((cell.values.get(column, Decimal("0")) for column in COMPONENT_TO_NODE), Decimal("0"))
        reported = cell.values.get("店铺利润", Decimal("0"))
        adjustment = reported - expected
        if abs(adjustment) > Decimal("0.05"):
            warnings.append(
                f"{cell.store_id} {cell.period} 保留历史结账调整 {adjustment}："
                f"可见组成项 {expected}，终态利润 {reported}"
            )
        receipt = abs(cell.values.get("交易收款", Decimal("0")))
        hard_limit = max(Decimal("10000"), receipt * Decimal("0.02"))
        if abs(adjustment) > hard_limit:
            errors.append(
                f"{cell.store_id} {cell.period} 历史调整 {adjustment} 超过门禁 {hard_limit}"
            )
    return ImportPlan(
        sorted(grouped.values(), key=lambda cell: (cell.period, cell.store_id)),
        sources,
        errors,
        warnings,
    )


def _node_row(node: StatementNode, value: Decimal) -> dict[str, Any]:
    return {
        "id": node.id,
        "name": node.name,
        "level": node.level,
        "value": float(value),
        "available": True,
        "display": node.display,
        "missing_sources": [],
        "is_total": node.is_total,
        "drillable": False,
    }


def payload(cell: LegacyCell, model: Model, model_revision: str) -> dict[str, Any]:
    values: dict[str, Decimal] = {node_id: Decimal("0") for node_id in ZERO_NODES}
    for column, node_id in COMPONENT_TO_NODE.items():
        values[node_id] = cell.values.get(column, Decimal("0"))
    values["g_revenue"] = values["n_receipt"] + values["n_refund"]
    values["g_platform"] = sum(
        (values[node] for node in (
            "n_software", "n_logistics", "n_cross_border", "n_compensation",
            "n_insurance", "n_marketing", "n_expense",
        )), Decimal("0"),
    )
    values["g_goods"] = values["n_goods"] + values["n_dropship"] + values["n_reshipment"]
    values["g_fulfillment"] = values["n_freight"]
    values["g_promotion"] = values["n_ad"] + values["n_small_payment"] + values["n_brushing"]
    values["gross_profit"] = cell.values.get("店铺毛利") or (
        values["n_receipt"] + values["n_goods"]
    )
    values["net_profit"] = cell.values["店铺利润"]
    values["net_margin"] = (
        values["net_profit"] / values["n_receipt"] if values["n_receipt"] else Decimal("0")
    )
    store = model.store(cell.store_id)
    statement = [_node_row(node, values.get(node.id, Decimal("0"))) for node in model.statement]
    visible_total = sum(
        (cell.values.get(column, Decimal("0")) for column in COMPONENT_TO_NODE),
        Decimal("0"),
    )
    adjustment = values["net_profit"] - visible_total
    if abs(adjustment) > Decimal("0.005"):
        statement.extend([
            {
                "id": "g_legacy_adjustment", "name": "历史结账调整", "level": 1,
                "value": float(adjustment), "available": True, "display": "amount",
                "missing_sources": [], "is_total": False, "drillable": False,
            },
            {
                "id": "n_legacy_adjustment", "name": "旧口径未显式列示差额", "level": 2,
                "value": float(adjustment), "available": True, "display": "amount",
                "missing_sources": [], "is_total": False, "drillable": False,
            },
        ])
    return {
        "store": store.name,
        "store_id": store.id,
        "platform": store.platform,
        "entity": store.entity,
        "period": cell.period,
        "can_close": True,
        "statement": statement,
        "findings": [{
            "id": "legacy_final_archive",
            "name": "历史终态归档",
            "passed": True,
            "blocking": False,
            "message": "该账期在台账系统上线前已完成，按只读终态结果映射并冻结。",
            "head": "历史终态结果已归档",
            "lines": [],
        }],
        "sources": [{
            "id": "legacy_final_summary",
            "name": "历史终态结果",
            "arrived": True,
            "reason": "只读归档，不参与后续账期自动计算",
        }],
        "missing_sources": [],
        "quality": [],
        "unclassified": [],
        "unlinked_total": 0.0,
        "unlinked_buckets": [],
        "rows": len(cell.row_numbers),
        "archive": {
            "kind": "legacy_final_summary",
            "read_only": True,
            "source_path": str(cell.source),
            "source_sha256": cell.source_sha256,
            "sheet": cell.sheet,
            "row_numbers": cell.row_numbers,
            "legacy_names": sorted(cell.legacy_names),
            "model_revision": model_revision,
            "visible_component_total": float(visible_total),
            "reported_profit": float(values["net_profit"]),
            "legacy_adjustment": float(adjustment),
            "mapped_at": _now(),
        },
    }


def archive_source(source: Path, sha: str, nas_root: Path) -> str:
    target = nas_root / "90_历史版本" / sha[:2] / sha / "payload"
    if target.exists():
        if _hash(target) != sha:
            raise ValueError(f"NAS历史库中的 {sha} 内容不匹配")
        return "existing"
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_name(f"payload.{source.stat().st_size}.part")
    shutil.copy2(source, temporary)
    if _hash(temporary) != sha:
        temporary.unlink(missing_ok=True)
        raise ValueError(f"历史文件复制后哈希不匹配：{source}")
    temporary.replace(target)
    return "copied"


def backup_database(workspace: Workspace, backup_dir: Path) -> Path:
    backup_dir.mkdir(parents=True, exist_ok=True)
    target = backup_dir / f"workspace-before-historical-close-{datetime.now():%Y%m%d-%H%M%S}.db"
    connection = sqlite3.connect(target)
    try:
        workspace.conn.backup(connection)
    finally:
        connection.close()
    return target


def apply_plan(
    workspace: Workspace,
    model: Model,
    model_revision: str,
    plan: ImportPlan,
    *,
    nas_root: Path,
    backup_dir: Path,
    cutoff: str = CUTOFF,
) -> dict[str, Any]:
    if plan.errors:
        raise WorkspaceError("历史终态计划有错误，禁止写入")
    if cutoff != CUTOFF:
        raise WorkspaceError(f"本迁移只允许截止 {CUTOFF}")
    if not nas_root.is_dir():
        raise WorkspaceError(f"NAS根目录不可达：{nas_root}")
    backup = backup_database(workspace, backup_dir)
    archive_counts = defaultdict(int)
    for source, sha in plan.sources.items():
        archive_counts[archive_source(source, sha, nas_root)] += 1

    imported = skipped = policy_closed = 0
    manifest_cells: list[dict[str, Any]] = []
    for cell in plan.cells:
        if cell.period > cutoff:
            raise WorkspaceError(f"越过截止期：{cell.store_id} {cell.period}")
        body = payload(cell, model, model_revision)
        current = workspace.state(cell.store_id, cell.period)
        if current and current.state == CLOSED:
            archive = (current.result or {}).get("archive") or {}
            if archive.get("source_sha256") == cell.source_sha256:
                skipped += 1
                continue
            raise WorkspaceError(f"{cell.store_id} {cell.period} 已由另一版本冻结")
        fingerprint = hashlib.sha256(
            json.dumps(
                [cell.source_sha256, cell.store_id, cell.period, model_revision, engine_version()],
                ensure_ascii=False,
            ).encode("utf-8")
        ).hexdigest()
        run_id = workspace.record(
            cell.store_id,
            cell.period,
            body,
            [cell.source_sha256],
            evidence_ready=True,
            model_revision=model_revision,
            input_fingerprint=fingerprint,
        )
        state = workspace.close_period(
            cell.store_id,
            cell.period,
            by=OPERATOR,
            note=f"台账上线前已完成；只读归档 {cell.source.name} {cell.sheet}",
        )
        imported += 1
        manifest_cells.append({
            "store_id": cell.store_id,
            "period": cell.period,
            "run_id": run_id,
            "state": state.state,
            "source_sha256": cell.source_sha256,
            "sheet": cell.sheet,
            "rows": cell.row_numbers,
        })

    # Any remaining current-workspace row inside the cutoff has no compatible consolidated final
    # report. Freeze its exact existing evidence under the user's explicit cutoff policy; do not
    # rewrite its result or pretend it passed current checks. This also covers a store that only
    # started appearing in the structured summary later in 2025.
    for state in workspace.overview():
        if not re.fullmatch(r"\d{4}-\d{2}", state.period or ""):
            continue
        if state.period > cutoff or state.state == CLOSED:
            continue
        workspace.close_historical_period(
            state.store_id,
            state.period,
            cutoff=cutoff,
            by=OPERATOR,
            note="按业务确认的历史截止政策冻结现有证据快照（未伪装为当前模型复算）",
            reference=f"所有 {cutoff} 及以前账期默认已结账",
        )
        policy_closed += 1

    manifest = {
        "schema": 1,
        "created_at": _now(),
        "cutoff": cutoff,
        "engine": engine_version(),
        "model_revision": model_revision,
        "backup": str(backup),
        "sources": [
            {"path": str(path), "sha256": sha, "archive": f"90_历史版本/{sha[:2]}/{sha}/payload"}
            for path, sha in plan.sources.items()
        ],
        "counts": {
            "imported": imported,
            "skipped": skipped,
            "policy_closed": policy_closed,
            "archive_copied": archive_counts["copied"],
            "archive_existing": archive_counts["existing"],
        },
        "cells": manifest_cells,
    }
    target = nas_root / "99_系统" / "migrations" / f"historical-close-{datetime.now():%Y%m%d-%H%M%S}.json"
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_suffix(".json.part")
    temporary.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    temporary.replace(target)
    manifest["manifest"] = str(target)
    return manifest


def summarize(plan: ImportPlan) -> dict[str, Any]:
    by_period: dict[str, int] = defaultdict(int)
    for cell in plan.cells:
        by_period[cell.period] += 1
    return {
        "cutoff": CUTOFF,
        "cells": len(plan.cells),
        "stores": len({cell.store_id for cell in plan.cells}),
        "periods": dict(sorted(by_period.items())),
        "sources": [{"path": str(path), "sha256": sha} for path, sha in plan.sources.items()],
        "errors": plan.errors,
        "warning_count": len(plan.warnings),
        "warnings": plan.warnings[:30],
    }


__all__ = [
    "apply_plan", "build_plan", "CUTOFF", "ImportPlan", "LEGACY_STORE_IDS",
    "payload", "summarize",
]
