from __future__ import annotations

import json
from copy import deepcopy
from pathlib import Path

from openpyxl import Workbook

from ledger.gaps import gaps
from ledger.historical_close import apply_plan, build_plan, canonical_legacy_name, payload
from ledger.model import load_model
from ledger.workspace import CLOSED, open_workspace


HEADERS = [
    "Name", "年月", "交易收款", "交易退款", "交易赔付", "软件服务费", "营销费用",
    "发货运费", "订单成本", "补发成本", "本金佣金", "代购代发", "店铺毛利",
    "广告费", "店铺利润",
]


def _values(name: str, period: str = "202501") -> list[object]:
    components = [1000, -50, -10, -40, -20, -80, -300, -5, -15, -30, -100]
    profit = sum(components)
    return [
        name, period, *components[:2], components[2], components[3], components[4],
        components[5], components[6], components[7], components[8], components[9], 700,
        components[10], profit,
    ]


def _source_2025(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "店铺明细"
    sheet.append(HEADERS)
    sheet.append(_values("美食专家"))
    workbook.save(path)


def _source_2026(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    headers = [column for column in HEADERS if column not in {"年月", "店铺毛利"}]
    for month in range(1, 6):
        sheet = workbook.create_sheet(f"{month}月")
        sheet.append(headers)
        row = dict(zip(HEADERS, _values("喜必顺旗舰店", f"2026{month:02d}"), strict=True))
        sheet.append([row[column] for column in headers])
    workbook.save(path)


def test_dated_legacy_store_name_is_collapsed() -> None:
    assert canonical_legacy_name("浅花涧旗舰店4.1-4.14") == "浅花涧旗舰店"
    assert canonical_legacy_name("#N/A") == ""


def test_plan_maps_legacy_reports_to_current_model(tmp_path: Path) -> None:
    source_2025 = tmp_path / "2025.xlsx"
    source_2026 = tmp_path / "2026.xlsx"
    _source_2025(source_2025)
    _source_2026(source_2026)
    model = load_model(Path(__file__).parents[2] / "models" / "cn-ecommerce")

    plan = build_plan(source_2025, source_2026, model)

    assert plan.errors == []
    assert len(plan.cells) == 6
    assert {(cell.store_id, cell.period) for cell in plan.cells} == {
        ("taobao_mt8egr48", "2025-01"),
        *(("taobao_xibishun", f"2026-{month:02d}") for month in range(1, 6)),
    }
    body = payload(plan.cells[0], model, "model-revision")
    assert body["archive"]["evidence_schema"] == 2
    assert len(body["findings"]) == 4
    assert len(body["sources"]) == 3
    assert len(body["archive"]["quality_checks"]) == 4
    assert body["archive"]["unlinked_evidence"]["status"] == "not_available"


def test_historical_gaps_only_expose_explicit_review_items(tmp_path: Path) -> None:
    source_2025 = tmp_path / "2025.xlsx"
    source_2026 = tmp_path / "2026.xlsx"
    _source_2025(source_2025)
    _source_2026(source_2026)
    model = load_model(Path(__file__).parents[2] / "models" / "cn-ecommerce")
    cell = build_plan(source_2025, source_2026, model).cells[0]
    body = payload(cell, model, "model-revision")
    body["archive"]["review_items"] = [{
        "kind": "historical_adjustment", "severity": "info", "title": "历史调整",
        "detail": "显式列示", "amount": -12.34,
    }]
    body["missing_sources"] = ["不应按当前门禁报告"]

    rows = gaps(body, model)

    assert rows == [{
        "kind": "historical_adjustment", "severity": "info", "title": "历史调整",
        "detail": "显式列示", "node": "", "metric": "", "source": "",
        "amount": -12.34, "only": "counted",
    }]


def test_apply_archives_sources_and_preserves_june_state(tmp_path: Path) -> None:
    source_2025 = tmp_path / "2025.xlsx"
    source_2026 = tmp_path / "2026.xlsx"
    _source_2025(source_2025)
    _source_2026(source_2026)
    model_root = Path(__file__).parents[2] / "models" / "cn-ecommerce"
    model = load_model(model_root)
    plan = build_plan(source_2025, source_2026, model)
    workspace = open_workspace(tmp_path / "workspace")
    nas = tmp_path / "nas"
    nas.mkdir()

    old = {"can_close": False, "findings": [], "missing_sources": [], "statement": []}
    workspace.record("taobao_mt8egr48", "2024-12", old, ["old"], evidence_ready=True)
    workspace.record("douyin_mszr2dhn", "2025-02", old, ["old-2025"], evidence_ready=True)
    june = {"can_close": True, "findings": [], "missing_sources": [], "statement": []}
    workspace.record("taobao_xibishun", "2026-06", june, ["june"], evidence_ready=True)
    workspace.close_period("taobao_xibishun", "2026-06", by="test")
    first = plan.cells[0]
    old_body = payload(first, model, "old-model-revision")
    old_body = deepcopy(old_body)
    old_body["archive"].pop("evidence_schema")
    old_body["archive"].pop("quality_checks")
    old_body["archive"].pop("unlinked_evidence")
    old_body["archive"].pop("review_items")
    old_body["findings"] = old_body["findings"][:1]
    old_body["sources"] = old_body["sources"][:1]
    old_run = workspace.record(
        first.store_id, first.period, old_body, [first.source_sha256], evidence_ready=True,
    )
    workspace.close_period(first.store_id, first.period, by="old-import")

    result = apply_plan(
        workspace,
        model,
        "model-revision",
        plan,
        nas_root=nas,
        backup_dir=tmp_path / "backups",
    )

    assert result["counts"]["imported"] == 5
    assert result["counts"]["enriched"] == 1
    assert result["counts"]["policy_closed"] == 2
    assert workspace.state("taobao_mt8egr48", "2024-12").state == CLOSED
    assert workspace.state("douyin_mszr2dhn", "2025-02").state == CLOSED
    assert workspace.state("taobao_mt8egr48", "2025-01").state == CLOSED
    enriched = workspace.state(first.store_id, first.period)
    assert enriched.run_id != old_run
    assert enriched.result["archive"]["evidence_schema"] == 2
    assert len(workspace.history(first.store_id, first.period)) == 2
    assert workspace.state("taobao_xibishun", "2026-05").state == CLOSED
    assert workspace.state("taobao_xibishun", "2026-06").state == CLOSED
    archived = list((nas / "90_历史版本").glob("*/*/payload"))
    assert len(archived) == 2
    manifest = Path(result["manifest"])
    assert json.loads(manifest.read_text(encoding="utf-8"))["cutoff"] == "2026-05"
    workspace.close()
